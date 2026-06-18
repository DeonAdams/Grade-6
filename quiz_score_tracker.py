"""
Fair Trade Adventure Quiz — Score Tracker
==========================================
Run this script to:
  1. Serve the quiz at http://localhost:5050
  2. Automatically log every completed section's score
     to an Excel file at C:\\Users\\01483051\\Jordan\\Quiz_Scores.xlsx

Usage:
  python quiz_score_tracker.py

Then open http://localhost:5050 in your browser.
Press Ctrl+C to stop the server.
"""

import json
import os
from datetime import datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────────
PORT = 5050
QUIZ_HTML = Path(__file__).parent / "Fair_Trade_Adventure_Quiz_tracked.html"
EXCEL_DIR = Path(r"C:\Users\01483051\Jordan")
EXCEL_FILE = EXCEL_DIR / "Quiz_Scores.xlsx"

# Fallback: if the Windows path doesn't exist (e.g. running on another OS),
# save next to this script instead.
if not EXCEL_DIR.exists():
    EXCEL_DIR = Path(__file__).parent
    EXCEL_FILE = EXCEL_DIR / "Quiz_Scores.xlsx"
    print(f"⚠  Target directory not found — saving to {EXCEL_FILE} instead.")


# ── Excel helpers ───────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1A5276")   # dark blue
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT     = Font(name="Arial", size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
THIN          = Side(style="thin", color="BFC9CA")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ROW_FILL_ODD  = PatternFill("solid", fgColor="EBF5FB")   # light blue stripe
ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")


def _col_widths():
    return {"A": 18, "B": 14, "C": 30, "D": 14}


def create_workbook():
    """Create a brand-new scores workbook with formatted headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz Scores"

    headers = ["Date", "Attempt No.", "Section Completed", "Score"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    ws.row_dimensions[1].height = 22
    for letter, width in _col_widths().items():
        ws.column_dimensions[letter].width = width

    wb.save(EXCEL_FILE)
    print(f"✅ Created new workbook: {EXCEL_FILE}")
    return wb


def load_or_create_workbook():
    if EXCEL_FILE.exists():
        return openpyxl.load_workbook(EXCEL_FILE)
    return create_workbook()


def next_attempt_number(ws):
    """Return the next attempt number (max existing + 1)."""
    max_attempt = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            try:
                max_attempt = max(max_attempt, int(row[1]))
            except (ValueError, TypeError):
                pass
    return max_attempt + 1


def append_score(section: str, points: int):
    """Append one score row to the Excel file."""
    wb = load_or_create_workbook()
    ws = wb.active

    attempt = next_attempt_number(ws)
    row_num = ws.max_row + 1
    fill = ROW_FILL_ODD if attempt % 2 == 1 else ROW_FILL_EVEN

    values = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        attempt,
        section,
        points,
    ]
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.font      = DATA_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        cell.fill      = fill

    wb.save(EXCEL_FILE)
    print(f"📊 Logged  Attempt {attempt} | {section} | Score: {points}  → {EXCEL_FILE}")


# ── HTTP server ─────────────────────────────────────────────────────────────────
class QuizHandler(BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        # Suppress default server noise; we print our own messages.
        pass

    def _send(self, code, content_type, body: bytes):
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", len(body))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        """CORS pre-flight."""
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, GET, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            try:
                html = QUIZ_HTML.read_bytes()
                self._send(200, "text/html; charset=utf-8", html)
            except FileNotFoundError:
                body = b"Quiz HTML not found. Make sure Fair_Trade_Adventure_Quiz_tracked.html is in the same folder."
                self._send(404, "text/plain", body)
        else:
            self._send(404, "text/plain", b"Not found")

    def do_POST(self):
        if self.path == "/log_score":
            length = int(self.headers.get("Content-Length", 0))
            raw = self.rfile.read(length)
            try:
                data = json.loads(raw)
                section = str(data.get("section", "Unknown"))
                points  = int(data.get("score", 0))
                append_score(section, points)
                self._send(200, "application/json", b'{"status":"ok"}')
            except Exception as exc:
                print(f"❌ Error logging score: {exc}")
                self._send(500, "application/json", b'{"status":"error"}')
        else:
            self._send(404, "text/plain", b"Not found")


# ── Entry point ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Ensure workbook exists before first request
    load_or_create_workbook()

    server = HTTPServer(("localhost", PORT), QuizHandler)
    print("=" * 60)
    print("  🌾 Fair Trade Adventure Quiz — Score Tracker")
    print("=" * 60)
    print(f"  Quiz URL  : http://localhost:{PORT}")
    print(f"  Excel file: {EXCEL_FILE}")
    print("  Press Ctrl+C to stop.")
    print("=" * 60)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n🛑 Server stopped.")
